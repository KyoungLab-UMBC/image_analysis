import openpyxl
from openpyxl.styles import Border, Side
from pathlib import Path

def get_existing_ids(sheet):
    """
    Finds all columns with 'Image Name' as the header (row 1) and 
    collects a set of all non-empty values from row 2 downwards in those columns.
    """
    ids = set()
    image_name_cols = []
    
    # Check the first row for 'Image Name' headers
    for cell in sheet[1]:
        if cell.value == "Image Name":
            image_name_cols.append(cell.column)
            
    for col_idx in image_name_cols:
        # Iterate through the rows in this specific column
        for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            if row[0] is not None:
                ids.add(row[0])
    return ids

def parse_image_name_for_sorting(file_path):
    """
    Extracts the image name from the file and formats it as a sortable string.
    Expected format: {yyyymmdd}-{a}-{b}-{folder name}
    Pads {a} and {b} with zeros so string sorting seamlessly matches numeric sorting.
    """
    try:
        # read_only=True makes scanning the files extremely fast
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if "Summary_num" in wb.sheetnames:
            ws = wb["Summary_num"]
            # Find the first 'Image Name' column
            for cell in ws[1]:
                if cell.value == "Image Name":
                    # Iterate to find the first non-empty image name
                    for row in ws.iter_rows(min_row=2, min_col=cell.column, max_col=cell.column, values_only=True):
                        if row[0]:
                            val = str(row[0]).strip()
                            parts = val.split('-')
                            if len(parts) >= 3:
                                p0 = parts[0]
                                p1 = parts[1].zfill(5)  # zero-pad (e.g., '1' -> '00001')
                                p2 = parts[2].zfill(5)  # zero-pad
                                return f"{p0}-{p1}-{p2}"
                            return val
    except Exception:
        pass
    return "99999999" # Push files with missing/invalid structures to the end

def process_and_merge_files(file_a_path_str, folder_b_path_str):
    try:
        # Convert strings to Path objects
        file_a_path = Path(file_a_path_str)
        folder_b_path = Path(folder_b_path_str)

        # 1. Load File A (The Master File)
        print(f"Loading Master File A: {file_a_path.name}...")
        if not file_a_path.exists():
            print(f"Error: File A not found at {file_a_path}")
            return

        wb_a = openpyxl.load_workbook(file_a_path)
        
        # Prepare the duplicate checker set from File A's "Summary_num" sheet
        existing_ids = set()
        if "Summary_num" in wb_a.sheetnames:
            existing_ids = get_existing_ids(wb_a["Summary_num"])
            print(f"Loaded {len(existing_ids)} existing IDs from 'Summary_num' sheet for duplicate checking.")
        else:
            print("Warning: 'Summary_num' sheet not found in File A. Duplicate checking might be limited.")

        # 2. Search for Excel files in Folder B (and subfolders)
        print(f"Searching for files starting with 'Cell' in {folder_b_path}...")
        found_files = list(folder_b_path.rglob("Cell*.xlsx"))
        
        if not found_files:
            print("No files starting with 'Cell' found.")
            return

        # --- SORTING ---
        print(f"Found {len(found_files)} files. Sorting by yyyymmdd-a-b sequence...")
        found_files.sort(key=parse_image_name_for_sorting)

        print("Starting append process...")
        files_processed_count = 0
        thin_bottom = Side(border_style="thin", color="000000")

        # 3. Iterate through each found file
        for file_b_path in found_files:
            print(f"\nChecking file: {file_b_path.name}")
            
            try:
                wb_b = openpyxl.load_workbook(file_b_path, data_only=True)
                
                # --- Duplicate Check ---
                skip_file = False
                if "Summary_num" in wb_b.sheetnames:
                    b_ids = get_existing_ids(wb_b["Summary_num"])
                    if not b_ids.isdisjoint(existing_ids):
                        print(f"  -> SKIP: IDs in {file_b_path.name} already exist in File A.")
                        skip_file = True
                
                if skip_file:
                    wb_b.close()
                    continue

                # --- Merge & Copy Logic ---
                print(f"  -> Processing sheets from {file_b_path.name}")
                
                for sheet_name in wb_b.sheetnames:
                    ws_b = wb_b[sheet_name]

                    # 1. Sheet exists in File A -> Run your existing merge logic
                    if sheet_name in wb_a.sheetnames:
                        ws_a = wb_a[sheet_name]
                        max_col_b = ws_b.max_column
                        max_row_b = ws_b.max_row

                        if max_row_b < 2:
                            continue

                        # Process vertically, column by column to handle unique positioning/blanks
                        for col_idx in range(1, max_col_b + 1):
                            header_b = ws_b.cell(row=1, column=col_idx).value
                            
                            # Extract the column's actual content from row 2 down
                            col_data = []
                            for r in range(2, max_row_b + 1):
                                col_data.append(ws_b.cell(row=r, column=col_idx).value)
                                
                            # Strip trailing trailing empty cells
                            while col_data and col_data[-1] is None:
                                col_data.pop()
                                
                            # If the entire column is empty (no header + no data), skip it
                            if not col_data and header_b is None:
                                continue
                                
                            # Handle the Header in Master file
                            header_a = ws_a.cell(row=1, column=col_idx).value
                            if header_a is None and header_b is not None:
                                ws_a.cell(row=1, column=col_idx).value = header_b
                                
                            # Find the bottom-most used row strictly in THIS column in Target
                            search_start = max(ws_a.max_row, 1)
                            max_row_a_col = 0
                            for r in range(search_start, 0, -1):
                                if ws_a.cell(row=r, column=col_idx).value is not None:
                                    max_row_a_col = r
                                    break
                                    
                            if not col_data:
                                continue
                                
                            # Paste the data cleanly beneath the last record
                            start_row_a = max_row_a_col + 1
                            for i, val in enumerate(col_data):
                                ws_a.cell(row=start_row_a + i, column=col_idx).value = val
                                
                            # Add ONLY a bottom border to the last pasted cell in this group
                            last_row_added = start_row_a + len(col_data) - 1
                            target_cell = ws_a.cell(row=last_row_added, column=col_idx)
                            
                            # Preserve existing side/top borders, add bottom border
                            old_border = target_cell.border
                            target_cell.border = Border(
                                left=old_border.left, right=old_border.right,
                                top=old_border.top, bottom=thin_bottom,
                                diagonal=old_border.diagonal, diagonal_direction=old_border.diagonal_direction,
                                outline=old_border.outline, vertical=old_border.vertical, horizontal=old_border.horizontal
                            )
                            
                            # Live-update existing IDs so duplicates within the run don't slip through
                            if sheet_name == "Summary_num" and header_b == "Image Name":
                                for val in col_data:
                                    if val is not None:
                                        existing_ids.add(val)

                    # 2. Sheet does NOT exist in File A -> Copy entire sheet at the same position
                    else:
                        print(f"  -> Copying new sheet: '{sheet_name}'")
                        
                        # Find the index position of the sheet in File B
                        b_idx = wb_b.sheetnames.index(sheet_name)
                        
                        # Create the sheet in File A at the exact same index
                        ws_a = wb_a.create_sheet(title=sheet_name, index=b_idx)
                        
                        # Copy all rows and values directly over
                        for row in ws_b.iter_rows(values_only=True):
                            ws_a.append(row)
                
                wb_b.close()
                files_processed_count += 1
                
            except Exception as e:
                print(f"  -> Error reading {file_b_path.name}: {e}")

        # 4. Save File A (Only once at the end)
        print(f"\nSaving updated File A to {file_a_path}...")
        wb_a.save(file_a_path)
        print(f"Success! Processed {files_processed_count} new files.")

    except Exception as e:
        print(f"An error occurred: {e}")


# --- execution block ---
if __name__ == "__main__":
    # 1. Path to your Master File (File A)
    file_a = r'F:\Weekly progression\Summary\Pyruvate\Pyruvate level in Glucosome radius analysis 180 mM 37C 20260407.xlsx'
    
    # 2. Path to the Folder containing Cell files (Folder B)
    folder_b = r'F:\20240710 PFKL-mCherry_PyronicSF_MitotrackerDeepred - High Salt Conc - 37 degree - WideField\Plate 2 - 180 mM'
    
    process_and_merge_files(file_a, folder_b)
import openpyxl
import os
import re

def reorganize_excel_data(file_path, data_sheet_name="Coloc Sm Att_pc", data_column_index=10):
    """
    Reads summary counts and redistributes data from a specified target sheet and column.
    
    Parameters:
    - file_path: Path to the excel file.
    - data_sheet_name: The name of the sheet to read the data from.
    - data_column_index: The number of the column to read (e.g., 1 for A, 2 for B, 10 for J).
    """
    # 1. Load the existing workbook
    print(f"Loading workbook: {file_path}")
    try:
        # data_only=True reads the actual values, not the formulas
        wb = openpyxl.load_workbook(file_path, data_only=True) 
    except FileNotFoundError:
        print("Error: The specified file was not found.")
        return

    # Ensure the required sheets exist
    if "Summary_num" not in wb.sheetnames:
        print("Error: The 'Summary_num' sheet is not in the workbook.")
        return
        
    if data_sheet_name not in wb.sheetnames:
        print(f"Error: The sheet '{data_sheet_name}' is not in the workbook.")
        return

    ws_summary = wb["Summary_num"]
    ws_data = wb[data_sheet_name]

    # 2. Read the "Summary_num" sheet, B2:Bn (extracting the numbers)
    counts = []
    # Loop through column B (column index 2), starting at row 2
    for row in range(2, ws_summary.max_row + 1):
        cell_value = ws_summary.cell(row=row, column=2).value
        # Stop reading if we hit an empty cell
        if cell_value is None:
            break
        # Ensure it's a number and add it to our list
        if isinstance(cell_value, (int, float)):
            counts.append(int(cell_value))

    print(f"Found {len(counts)} numbers in 'Summary_num' Column B.")

    # 3. Read the dynamic data sheet and dynamic column
    # Get the header (row 1) of the selected column for the filename
    header_value = str(ws_data.cell(row=1, column=data_column_index).value)
    
    # Sanitize the header to remove characters that are invalid in file names
    safe_header = re.sub(r'[\\/*?:"<>|]', "", header_value)

    # Extract all data from row 2 to the bottom of the sheet for the specific column
    col_data = []
    for row in range(2, ws_data.max_row + 1):
        col_data.append(ws_data.cell(row=row, column=data_column_index).value)

    # 4. Create the new Excel file and map the data
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "Reorganized Data"

    current_data_index = 0
    
    # Loop through our list of numbers (counts)
    for col_idx, count in enumerate(counts, start=1):
        # Slice the column data to get the exact number of cells needed
        chunk = col_data[current_data_index : current_data_index + count]
        
        # Paste this chunk into the new worksheet, column by column
        for row_idx, value in enumerate(chunk, start=1):
            new_ws.cell(row=row_idx, column=col_idx, value=value)
            
        # Move our starting index forward by the number of cells we just copied
        current_data_index += count

    # Determine the save location (same folder as the original file)
    base_dir = os.path.dirname(os.path.abspath(file_path))
    
    # Output filename now dynamically updates based on the sheet name AND the column header
    new_filename = f"{data_sheet_name} + {safe_header}.xlsx"
    new_filepath = os.path.join(base_dir, new_filename)

    # Save the new workbook
    new_wb.save(new_filepath)
    print(f"Success! New file saved at: {new_filepath}")


# --- How to run the script ---
if __name__ == "__main__":
    # Replace with the actual path to your Excel file
    my_file_path = r"F:\Weekly progression\Summary\Pyruvate\Pyruvate level in Glucosome radius analysis 140 mM 37C 20260407.xlsx" 
    
    # Variables you can easily change at any time:
    target_sheet = "Coloc Sm Att_pc"
    target_column_number = 5  # 10 means Column J, 1 means Column A, 2 means Column B, etc.
    
    # Run the function with your variables
    reorganize_excel_data(
        file_path=my_file_path, 
        data_sheet_name=target_sheet, 
        data_column_index=target_column_number
    )